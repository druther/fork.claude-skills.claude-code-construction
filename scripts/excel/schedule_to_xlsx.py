#!/usr/bin/env python3
"""Convert extracted schedule data (JSON) to a formatted Excel workbook."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import safe_output_path

def schedule_to_xlsx(data_path, output="schedule.xlsx", schedule_type="generic",
                     project="", sheet=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    with open(data_path) as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = schedule_type.replace("_", " ").title()

    # Header
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Project info rows
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Project: {project}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Source: Sheet {sheet} | Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    # Determine columns from data
    if isinstance(data, list) and len(data) > 0:
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [list(row.values()) for row in data]
        elif isinstance(data[0], list):
            headers = data[0]
            rows = data[1:]
        else:
            headers = ["Value"]
            rows = [[v] for v in data]
    else:
        print("ERROR: Data must be a non-empty list")
        sys.exit(1)

    # Write headers
    start_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=str(header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Write data rows
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r, row in enumerate(rows, start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if (r - start_row) % 2 == 0:
                cell.fill = alt_fill

    # Auto-size columns (approximate)
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(start_row, start_row + len(rows) + 1)
        )
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = min(max_len + 4, 40)

    # Metadata tab
    meta = wb.create_sheet("Metadata")
    meta["A1"] = "Field"
    meta["B1"] = "Value"
    meta_data = [
        ("Project", project),
        ("Source Sheet", sheet),
        ("Schedule Type", schedule_type),
        ("Extraction Date", datetime.now().isoformat()),
        ("Row Count", len(rows)),
        ("Column Count", len(headers)),
        ("Tool", "construction-skills/schedule-extractor"),
    ]
    for r, (k, v) in enumerate(meta_data, 2):
        meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        meta.cell(row=r, column=2, value=str(v))

    out = safe_output_path(output)
    wb.save(str(out))
    print(f"OK: {out} ({len(rows)} rows, {len(headers)} columns)")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert schedule data to Excel")
    parser.add_argument("--data", required=True, help="JSON data file")
    parser.add_argument("--output", "-o", default="schedule.xlsx")
    parser.add_argument("--type", default="generic", help="Schedule type")
    parser.add_argument("--project", default="", help="Project name")
    parser.add_argument("--sheet", default="", help="Source sheet number")
    args = parser.parse_args()
    schedule_to_xlsx(args.data, args.output, args.type, args.project, args.sheet)
