#!/usr/bin/env python3
"""Convert submittal requirement data to a formatted submittal register Excel workbook."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from shared import safe_output_path

def submittal_log_to_xlsx(data_path, output="submittal_register.xlsx", project=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    with open(data_path) as f:
        items = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Submittal Register"
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"SUBMITTAL REGISTER — {project}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "Submittal #", "Spec Section", "Section Title", "Description",
        "Type", "Responsible Party", "Required Date", "Submitted Date",
        "Status", "Approved Date", "Notes"
    ]
    widths = [14, 14, 25, 35, 16, 20, 14, 14, 12, 14, 30]

    row = 4
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w

    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    for r, item in enumerate(items, row + 1):
        vals = [
            item.get("submittal_number", ""),
            item.get("spec_section", ""),
            item.get("section_title", ""),
            item.get("description", ""),
            item.get("type", ""),
            item.get("responsible_party", ""),
            "",  # Required date (user fills in)
            "",  # Submitted date
            "",  # Status
            "",  # Approved date
            item.get("notes", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True)
            if (r - row) % 2 == 0:
                cell.fill = alt_fill

    # Summary tab
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Submittal Register Summary"
    summary["A1"].font = Font(bold=True, size=12)
    summary["A3"] = f"Total submittals: {len(items)}"

    # Count by type
    types = {}
    for item in items:
        t = item.get("type", "Unknown")
        types[t] = types.get(t, 0) + 1
    r = 5
    summary["A5"] = "By Type:"
    summary["A5"].font = Font(bold=True)
    for t, count in sorted(types.items()):
        r += 1
        summary.cell(row=r, column=1, value=t)
        summary.cell(row=r, column=2, value=count)

    out = safe_output_path(output)
    wb.save(str(out))
    print(f"OK: {out} ({len(items)} submittals)")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert submittal data to Excel register")
    parser.add_argument("--data", required=True, help="JSON data file")
    parser.add_argument("--output", "-o", default="submittal_register.xlsx")
    parser.add_argument("--project", default="")
    args = parser.parse_args()
    submittal_log_to_xlsx(args.data, args.output, args.project)
